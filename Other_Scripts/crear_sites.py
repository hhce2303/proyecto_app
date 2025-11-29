import os
from openpyxl import Workbook, load_workbook

# Ruta del Excel de sites original
EXCEL_SITES_ORIGINAL = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Customer Sites     SIG Tools -    Cloud.xlsx"
SHEET_NAME = "Sheet1"

# Ruta donde se guardará el Excel de sites
EXCEL_SITES_DB = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\SitesDatabase.xlsx"

def read_sites():
    """Leer sites desde el Excel original"""
    if not os.path.exists(EXCEL_SITES_ORIGINAL):
        print("El archivo de sites original no existe.")
        return []

    wb = load_workbook(EXCEL_SITES_ORIGINAL, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"La hoja {SHEET_NAME} no existe.")
        return []

    ws = wb[SHEET_NAME]
    sites = []

    for row in ws.iter_rows(min_row=3, min_col=3, max_col=3, values_only=True):
        site_name = row[0]
        if site_name and str(site_name).strip():
            sites.append(str(site_name).strip())
    return sites

def create_sites_excel():
    """Crear un Excel nuevo con la lista de sites"""
    sites = read_sites()
    if not sites:
        print("No se encontraron sites para guardar.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Sites"

    # Encabezado
    ws.append(["Site Name"])

    # Agregar sites
    for s in sites:
        ws.append([s])

    wb.save(EXCEL_SITES_DB)
    print(f"Excel de sites creado correctamente en: {EXCEL_SITES_DB}")

if __name__ == "__main__":
    create_sites_excel()
