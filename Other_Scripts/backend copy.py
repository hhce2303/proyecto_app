import os
import csv
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Rutas de archivos ---
EXCEL_FILE = r"C:\Users\hcruz.SIG\OneDrive - SIG Systems, Inc\Desktop\proyecto_app\BaseDatos.xlsx"
USERS_FILE = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Usuarios.xlsx"
SITES_FILE = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\SitesDatabase.xlsx"
OPERATORS_BASE_PATH = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators"
CUSTOMER_SITES_ORIGINAL = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Customer Sites     SIG Tools -    Cloud.xlsx"
ACTIVITIES_FILE = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Actividades.xlsx"

# ---------------------------
# Timestamp global de login
# ---------------------------
LOGIN_TIMESTAMP = datetime.now()

# ---------------------------
# Inicialización de Excels
# ---------------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(["ID", "Site", "Actividad", "Cantidad", "Camera", "Descripción", "Usuario", "Estado", "Timestamp"])
        wb.save(EXCEL_FILE)

def init_users_excel():
    os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
    if os.path.exists(USERS_FILE):
        return
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    users = get_users_from_path(latest_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["Usuario", "FechaCreacion"])
    for user in users:
        ws.append([user, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(USERS_FILE)

def init_sites_excel():
    os.makedirs(os.path.dirname(SITES_FILE), exist_ok=True)
    if os.path.exists(SITES_FILE):
        return
    sites = read_sites_from_source()
    if not sites:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites"
    ws.append(["Site Name"])
    for s in sites:
        ws.append([s])
    wb.save(SITES_FILE)

# ---------------------------
# Funciones de usuarios
# ---------------------------
def get_latest_year_month_path(base_path):
    años = [d for d in os.listdir(base_path) if d.isdigit()]
    if not años:
        raise FileNotFoundError("No se encontraron carpetas de años.")
    latest_year = max(años)
    year_path = os.path.join(base_path, latest_year)
    meses = [d for d in os.listdir(year_path) if os.path.isdir(os.path.join(year_path, d))]
    meses_sorted = sorted(meses, key=lambda x: int(x.split(".")[0]))
    latest_month = meses_sorted[-1]
    return os.path.join(year_path, latest_month)

def get_users_from_path(path):
    return sorted([d for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))])

def update_users_excel():
    os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    detected_users = get_users_from_path(latest_path)

    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        ws.append(["Usuario", "FechaCreacion"])
        for u in detected_users:
            ws.append([u, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(USERS_FILE)
        return detected_users

    wb = load_workbook(USERS_FILE)
    ws = wb["Usuarios"]
    existing_users = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
    new_users = [u for u in detected_users if u not in existing_users]

    for user in new_users:
        ws.append([user, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    if new_users:
        wb.save(USERS_FILE)

    return detected_users

def get_users():
    return update_users_excel()

def get_user_folder(username):
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    user_folder = os.path.join(latest_path, username)
    os.makedirs(user_folder, exist_ok=True)
    return user_folder

def create_user(username):
    if not username or username.strip() == "":
        raise ValueError("El nombre de usuario no puede estar vacío")
    username = username.strip()
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    user_folder = os.path.join(latest_path, username)
    if os.path.exists(user_folder):
        raise FileExistsError(f"El usuario '{username}' ya existe en {latest_path}")
    os.makedirs(user_folder)

    if not os.path.exists(USERS_FILE):
        init_users_excel()

    wb = load_workbook(USERS_FILE)
    ws = wb["Usuarios"]
    existing_users = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
    if username not in existing_users:
        ws.append([username, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(USERS_FILE)
    return username

# ---------------------------
# Funciones de sites
# ---------------------------
def read_sites_from_source():
    if not os.path.exists(CUSTOMER_SITES_ORIGINAL):
        return []
    wb = load_workbook(CUSTOMER_SITES_ORIGINAL, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    sites = []
    for row in ws.iter_rows(min_row=3, min_col=3, max_col=3, values_only=True):
        if row[0] and str(row[0]).strip():
            sites.append(str(row[0]).strip())
    return sites

def get_sites():
    init_sites_excel()
    wb = load_workbook(SITES_FILE)
    ws = wb["Sites"]
    return [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]

# ---------------------------
# Funciones de tickets
# ---------------------------
def add_record_custom(site, actividad, cantidad, camera, desc, user):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Tickets"]
    new_id = ws.max_row
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([new_id, site, actividad, cantidad, camera, desc, user, "Abierto", timestamp])
    wb.save(EXCEL_FILE)
    return f"Registro agregado con ID {new_id} a las {timestamp}"

def get_user_csv_path(user_name):
    global LOGIN_TIMESTAMP
    now = datetime.now()
    if now - LOGIN_TIMESTAMP > timedelta(hours=11):
        LOGIN_TIMESTAMP = now
    user_folder = get_user_folder(user_name)
    file_name = LOGIN_TIMESTAMP.strftime("%d-%m-%Y") + ".csv"
    return os.path.join(user_folder, file_name)

def add_record_user_log_full(site, actividad, cantidad, camera, desc, user):
    msg = add_record_custom(site, actividad, cantidad, camera, desc, user)
    csv_file = get_user_csv_path(user)
    file_exists = os.path.exists(csv_file)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [site, actividad, cantidad, camera, desc, user, timestamp]
    with open(csv_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["Site", "Actividad", "Cantidad", "Camera", "Descripcion", "Usuario", "Timestamp"])
        writer.writerow(row)
    return msg + f"\nEvento guardado en: {csv_file}"

def get_events(user_name):
    global LOGIN_TIMESTAMP

    now = datetime.now()
    # Reiniciar timestamp si ya pasó más de 11 horas
    if now - LOGIN_TIMESTAMP > timedelta(hours=11):
        LOGIN_TIMESTAMP = now

    # Construir ruta del CSV diario del usuario
    user_folder = get_user_folder(user_name)
    file_name = LOGIN_TIMESTAMP.strftime("%d-%m-%Y") + ".csv"
    file_path = os.path.join(user_folder, file_name)

    # Si no existe, devuelve lista vacía
    if not os.path.exists(file_path):
        return []

    # Leer CSV
    with open(file_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        return list(reader)

def generate_report(user_name):
    """Genera resumen por Site y Actividad para un usuario, desde su CSV"""
    user_folder = get_user_folder(user_name)
    from datetime import datetime
    file_name = datetime.now().strftime("%d-%m-%Y") + ".csv"
    file_path = os.path.join(user_folder, file_name)

    if not os.path.exists(file_path):
        return []  # Sin datos

    summary = {}
    import csv
    with open(file_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader, None)  # Saltar header
        for row in reader:
            key = (row[1], row[0])  # Site y Actividad
            summary[key] = summary.get(key, 0) + int(row[3])

    return [(site, act, total) for (site, act), total in summary.items()]

# ---------------------------
# Exportar eventos y borrar CSV
# ---------------------------
def export_events_to_excel(user_name):
    """Toma los CSV del usuario para la fecha de login, exporta a Excel y borra los CSV"""
    user_folder = get_user_folder(user_name)
    date_str = LOGIN_TIMESTAMP.strftime("%d-%m-%Y")

    # Buscar todos los CSV del día
    csv_files = [f for f in os.listdir(user_folder) if f.endswith(".csv") and f.startswith(date_str)]
    if not csv_files:
        raise FileNotFoundError(f"No se encontraron archivos CSV del día {date_str} en {user_folder}")

    events = []
    for csv_file in csv_files:
        with open(os.path.join(user_folder, csv_file), newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row in reader:
                events.append(row)

    if not events:
        raise ValueError("Los CSV están vacíos, no hay eventos para exportar.")

    # Crear Excel estético
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"

    headers = ["Site", "Actividad", "Cantidad", "Camera", "Descripcion", "Usuario", "Timestamp"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(col_name) + 2)

    for event in events:
        ws.append(event)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Guardar Excel final con sufijo de hora-minuto del login
    login_time = LOGIN_TIMESTAMP.strftime("%H-%M")
    export_name = f"{date_str}_{login_time}.xlsx"
    export_path = os.path.join(user_folder, export_name)
    wb.save(export_path)

    # ---------------------------
    # Borrar todos los CSV del día
    for csv_file in csv_files:
        os.remove(os.path.join(user_folder, csv_file))

    return export_path

# ---------------------------
# Funciones de actividades
# ---------------------------
def get_activities():
    if not os.path.exists(ACTIVITIES_FILE):
        return []
    try:
        wb = load_workbook(ACTIVITIES_FILE, data_only=True)
        if "Actividades" not in wb.sheetnames:
            return []
        ws = wb["Actividades"]
        return [cell.value for cell in ws['A'][1:] if cell.value]
    except:
        return []
