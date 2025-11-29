import os
import csv
import smtplib
import time
import ctypes
import threading
import tkinter as tk
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.text import MIMEText
from collections import defaultdict
from tkinter import messagebox




# --- Rutas de archivos ---
EXCEL_FILE = r"C:\Users\hcruz.SIG\OneDrive - SIG Systems, Inc\Desktop\proyecto_app\BaseDatos.xlsx"
USERS_FILE = r"\\192.168.7.12\\Data SIG\Central Station SLC-COLOMBIA\\1. Daily Logs - Operators\\DataBase\Usuarios.xlsx"
SITES_FILE = r"\\192.168.7.12\\Data SIG\Central Station SLC-COLOMBIA\\1. Daily Logs - Operators\\DataBase\SitesDatabase.xlsx"
OPERATORS_BASE_PATH = r"\\192.168.7.12\Data SIG\\Central Station SLC-COLOMBIA\\1. Daily Logs - Operators"
CUSTOMER_SITES_ORIGINAL = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Customer Sites     SIG Tools -    Cloud.xlsx"
ACTIVITIES_FILE = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\DataBase\Actividades.xlsx"
SESSION_CSV_FILES = {}
REPORTS_PATH = r"\\192.168.7.12\Data SIG\Central Station SLC-COLOMBIA\1. Daily Logs - Operators\Incidents Report"

# ---------------------------
# Timestamp global de login
# ---------------------------
LOGIN_TIMESTAMP = datetime.now()

# ---------------------------
# Inicialización de Excels
# ---------------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(["ID", "Site", "Actividad", "Cantidad", "Camera", "Descripción", "Usuario", "Estado", "Timestamp"])
        wb.save(EXCEL_FILE)

def init_users_excel():
    os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
    if os.path.exists(USERS_FILE):
        return
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    users = get_users_from_path(latest_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["Usuario", "FechaCreacion"])
    for user in users:
        ws.append([user, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(USERS_FILE)

def init_sites_excel():
    os.makedirs(os.path.dirname(SITES_FILE), exist_ok=True)
    if os.path.exists(SITES_FILE):
        return
    sites = read_sites_from_source()
    if not sites:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sites"
    ws.append(["Site Name"])
    for s in sites:
        ws.append([s])
    wb.save(SITES_FILE)

# ---------------------------
# Funciones de usuarios
# ---------------------------
def get_latest_year_month_path(base_path=OPERATORS_BASE_PATH):
    """Devuelve la ruta más reciente de año/mes o None si no existe"""
    if not os.path.exists(base_path):
        return None

    años = [d for d in os.listdir(base_path) if d.isdigit() and os.path.isdir(os.path.join(base_path, d))]
    if not años:
        return None
    latest_year = max(años)
    year_path = os.path.join(base_path, latest_year)

    meses = [d for d in os.listdir(year_path) if os.path.isdir(os.path.join(year_path, d))]
    if not meses:
        return None
    meses_sorted = sorted(meses, key=lambda x: int(x.split(".")[0]))
    latest_month = meses_sorted[-1]

    return os.path.join(year_path, latest_month)

def get_latest_year_month_filtered(base_path, filter_word):
    latest_path = get_latest_year_month_path(base_path)
    if not latest_path:
        return None

    for d in os.listdir(latest_path):
        full_path = os.path.join(latest_path, d)
        if os.path.isdir(full_path) and filter_word.lower() in d.lower():
            return full_path

    return latest_path


def get_users_from_path(path):
    """Devuelve lista de usuarios, vacía si path no existe"""
    if not path or not os.path.exists(path):
        return []
    return sorted([d for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))])


def get_user_folder(username):
    """Devuelve la carpeta del usuario, la crea si no existe"""
    latest_path = get_latest_year_month_path()
    if not latest_path:
        # Se puede crear la carpeta base si quieres
        print("⚠️ No hay carpeta de año/mes disponible")
        return None

    user_folder = os.path.join(latest_path, username)
    os.makedirs(user_folder, exist_ok=True)
    return user_folder


def get_users():
    """Devuelve lista de usuarios segura"""
    try:
        latest_path = get_latest_year_month_path()
        return get_users_from_path(latest_path)
    except Exception as e:
        print(f"Error al obtener usuarios: {e}")
        return []

def show_alert(title, message):
    ctypes.windll.user32.MessageBoxW(0, message, title, 0x40)  # 0x40 = icono de información

def get_user_folder(username, wait_time=10):
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    user_folder = os.path.join(latest_path, username)

    start = time.time()
    while not os.path.exists(user_folder):
        try:
            os.makedirs(user_folder, exist_ok=True)
        except Exception as e:
            show_alert("Error", f"Error creando carpeta:\n{e}")
            break

        if time.time() - start > wait_time:
            show_alert("Advertencia", f"La carpeta {user_folder} aún no está disponible después de {wait_time}s.\nPuede estar lento el acceso.")
            break
        time.sleep(1)  # espera activa

    if os.path.exists(user_folder):
        print("Éxito", f"Carpeta lista:\n{user_folder}")
    else:
        show_alert("Error", f"Carpeta no encontrada:\n{user_folder}")

    return user_folder

def create_user(username):
    if not username or username.strip() == "":
        raise ValueError("El nombre de usuario no puede estar vacío")
    username = username.strip()
    latest_path = get_latest_year_month_path(OPERATORS_BASE_PATH)
    user_folder = os.path.join(latest_path, username)
    if os.path.exists(user_folder):
        raise FileExistsError(f"El usuario '{username}' ya existe en {latest_path}")
    os.makedirs(user_folder)

    if not os.path.exists(USERS_FILE):
        init_users_excel()

    wb = load_workbook(USERS_FILE)
    ws = wb["Usuarios"]
    existing_users = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}
    if username not in existing_users:
        ws.append([username, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(USERS_FILE)
    return username

# ---------------------------
# Funciones de sites
# ---------------------------
def read_sites_from_source():
    if not os.path.exists(CUSTOMER_SITES_ORIGINAL):
        return []
    wb = load_workbook(CUSTOMER_SITES_ORIGINAL, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    sites = []
    for row in ws.iter_rows(min_row=3, min_col=3, max_col=3, values_only=True):
        if row[0] and str(row[0]).strip():
            sites.append(str(row[0]).strip())
    return sites

def init_sites_excel():
    """Crea el archivo de sites si no existe."""
    if not os.path.exists(SITES_FILE):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sites"
        ws.append(["Nombre"])
        wb.save(SITES_FILE)

def get_sites():
    """Lee los sitios desde el Excel sites.xlsx."""
    init_sites_excel()
    wb = load_workbook(SITES_FILE)
    ws = wb["Sites"]

    sites = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:  # evita filas vacías
            site_name = str(row[0]).strip()
            timezone = str(row[1]).strip() if len(row) > 1 and row[1] else "CT"  # default CT
            sites.append([site_name, timezone])
    wb.close()
    return sites

# ---------------------------
# Funciones de tickets
# ---------------------------
def add_record_custom(site, actividad, cantidad, camera, desc, user):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Tickets"]
    new_id = ws.max_row
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([new_id, site, actividad, cantidad, camera, desc, user, "Abierto", timestamp])
    wb.save(EXCEL_FILE)
    return print (f"Registro agregado con ID {new_id} a las {timestamp}")

def get_user_csv_path(user_name):
    global LOGIN_TIMESTAMP
    now = datetime.now()
    if now - LOGIN_TIMESTAMP > timedelta(hours=11):
        LOGIN_TIMESTAMP = now
    user_folder = get_user_folder(user_name)
    file_name = LOGIN_TIMESTAMP.strftime("%d-%m-%Y") + ".csv"
    return os.path.join(user_folder, file_name)

def add_record_user_log_full(site, activity, quantity, camera, desc, user, event_time):
    """
    Guarda un evento en un CSV individual (timestamp único) con las columnas:
    Timestamp, Usuario, Especial, Site, Actividad, Cantidad, Camera, Descripción, Registrado
    """
    user_folder = get_user_folder(user)
    os.makedirs(user_folder, exist_ok=True)

    # Nombre del archivo basado en la hora exacta del evento
    file_name = event_time.strftime("%Y-%m-%d_%H-%M-%S") + ".csv"
    file_path = os.path.join(user_folder, file_name)

    # Prefijos especiales
    especiales = ("AS", "DT", "HUD", "SCH", "KG", "LT", "PE", "WAG")
    es_especial = site.startswith(especiales)
    valor_especial = "✔" if es_especial else ""
    valor_registrado = "☐"  # inicialmente desmarcado

    # Encabezados completos
    fieldnames = ["Timestamp", "Usuario", "Especial", "Site",
                  "Actividad", "Cantidad", "Camera", "Descripción", "Registrado"]

    # Datos del evento
    row = {
        "Timestamp": event_time.strftime("%Y-%m-%d %H:%M:%S"),
        "Usuario": user,
        "Especial": valor_especial,
        "Site": site,
        "Actividad": activity,
        "Cantidad": quantity,
        "Camera": camera,
        "Descripción": desc,
        "Registrado": valor_registrado,
    }

    # Escribir CSV (un registro por archivo)
    with open(file_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow(row)

    # Registrar en la sesión, por usuario
    if user not in SESSION_CSV_FILES:
        SESSION_CSV_FILES[user] = []
    SESSION_CSV_FILES[user].append(file_path)

    return f"Evento guardado en {file_path}"

def get_events(user_name):
    """
    Devuelve todos los eventos del usuario para el mes actual.
    Lee todos los CSV existentes en la carpeta del usuario.
    """
    user_folder = get_user_folder(user_name)

    if not os.path.exists(user_folder):
        return []

    all_events = []
    headers_added = False

    # Listar todos los CSVs de la carpeta del usuario
    csv_files = [f for f in os.listdir(user_folder) if f.lower().endswith(".csv")]

    # Ordenar cronológicamente por nombre de archivo (timestamp)
    for file_name in sorted(csv_files):
        file_path = os.path.join(user_folder, file_name)
        try:
            with open(file_path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
                if not rows:
                    continue
                # Agregar header solo una vez
                if not headers_added:
                    all_events.append(rows[0])
                    headers_added = True
                # Agregar filas (sin repetir header)
                all_events.extend(rows[1:])
        except Exception as e:
            print(f"⚠️ No se pudo leer {file_path}: {e}")
            continue

    return all_events

def refresh_event_tree(user_name, event_tree):
    # Limpiar el Treeview
    for item in event_tree.get_children():
        event_tree.delete(item)

    # Obtener eventos actualizados
    all_events = get_events(user_name)

    if not all_events:
        return

    headers = all_events[0]
    rows = all_events[1:]

    # Configurar columnas (solo la primera vez)
    if not event_tree["columns"]:
        event_tree["columns"] = headers
        for col in headers:
            event_tree.heading(col, text=col)
            event_tree.column(col, width=100, anchor="center")

    # Insertar filas
    for row in rows:
        event_tree.insert("", tk.END, values=row)

    # Llamar nuevamente a esta función después de 5000 ms (5 segundos)
    event_tree.after(5000, lambda: refresh_event_tree(user_name, event_tree))

def save_report(site, report_type, desc, username):
    """Guarda un reporte CSV en la ruta compartida."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = os.path.join(REPORTS_PATH, f"{timestamp}_{report_type}_{username}.csv")

    row = {
        "Timestamp": timestamp,
        "Site": site,
        "ReportType": report_type,
        "Description": desc,
        "User": username
    }

    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=row.keys())
        writer.writeheader()
        writer.writerow(row)

    return filename

# ---------------------------
# Exportar eventos y borrar CSV
# ---------------------------
def export_events_to_excel(user_name):
    """
    Exporta a Excel todos los CSV del turno (fecha LOGIN_TIMESTAMP) del usuario,
    incluyendo los que no están en la sesión actual si el usuario acepta, y borra solo los CSV de la sesión.
    """
    user_folder = get_user_folder(user_name)
    date_str = LOGIN_TIMESTAMP.strftime("%Y-%m-%d")

    if not os.path.exists(user_folder):
        raise FileNotFoundError(f"No se encontró la carpeta del usuario: {user_folder}")

    # Buscar todos los CSV del día (turno)
    csv_files = [
        os.path.join(user_folder, f)
        for f in os.listdir(user_folder)
        if f.endswith(".csv")
    ]

    # Separar CSV del día del turno y CSV “extra”
    csv_files_turno = [f for f in csv_files if f.startswith(date_str)]
    csv_files_extra = [f for f in csv_files if f not in csv_files_turno]

    # Si hay CSV extra, preguntar al usuario si desea incluirlos
    if csv_files_extra:
        include_extra = messagebox.askyesno(
            "CSV detectados",
            f"Se encontraron {len(csv_files_extra)} CSV"
            "¿Desea incluirlos en la exportación?"
        )
        if include_extra:
            csv_files_turno.extend(csv_files_extra)

    if not csv_files_turno:
        raise FileNotFoundError(f"No se encontraron archivos CSV del día {date_str} en {user_folder}")

    events = []
    for csv_file in csv_files_turno:
        with open(csv_file, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            for row in reader:
                events.append(row)

    if not events:
        raise ValueError("No hay eventos para exportar.")

    # Crear Excel estético (igual que antes)
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    headers = ["Timestamp", "Usuario", "Especial", "Site", "Actividad", "Cantidad", "Camara", "Descripcion", ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(col_name) + 2)

    for event in events:
        ws.append(event)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Guardar Excel
    login_time = LOGIN_TIMESTAMP.strftime("%H-%M")
    export_name = f"{date_str}_{login_time}.xlsx"
    export_path = os.path.join(user_folder, export_name)
    wb.save(export_path)

    
    # ---------------------------
    # Borrar SOLO los CSV del turno, incluyendo fuera de la sesión
    for csv_file in csv_files_turno:
        if csv_file.lower().endswith(".csv"):  # <-- Esto evita borrar XLSX u otros
            try:
                os.remove(csv_file)
            except Exception as e:
                print(f"⚠️ No se pudo borrar {csv_file}: {e}")

    # Limpiar la sesión
    if user_name in SESSION_CSV_FILES:
        SESSION_CSV_FILES[user_name] = []

    return export_path


# ---------------------------
# Funciones de actividades
# ---------------------------
def get_activities():
    if not os.path.exists(ACTIVITIES_FILE):
        return []
    try:
        wb = load_workbook(ACTIVITIES_FILE, data_only=True)
        if "Actividades" not in wb.sheetnames:
            return []
        ws = wb["Actividades"]
        return [cell.value for cell in ws['A'][1:] if cell.value]
    except:
        return []
